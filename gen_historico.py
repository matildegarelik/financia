import openpyxl
from datetime import datetime

wb = openpyxl.load_workbook('mami_bills.xlsx', data_only=True)

UID = '402ad72e-fbb4-4316-aa69-669dd4296bda'

# ── helpers ────────────────────────────────────────────────────────────────

def esc(s):
    if s is None: return 'NULL'
    s = str(s).strip()
    if not s: return 'NULL'
    return "'" + s.replace("'", "''") + "'"

def fmt_date(d):
    if d is None: return 'NULL'
    if isinstance(d, datetime): return "'" + d.strftime('%Y-%m-%d') + "'"
    s = str(d).strip()
    return "'" + s + "'" if s else 'NULL'

def fmt_num(n):
    if n is None: return 'NULL'
    try: return str(round(float(n), 4))
    except: return 'NULL'

def norm_currency(c):
    if c is None: return 'ARS'
    return str(c).replace('$', '').strip().upper() or 'ARS'

# Normaliza nombres de cuentas del Excel a un nombre canónico interno
RAW_TO_CANONICAL = {
    'fiverr': 'Fiverr', 'fIverr': 'Fiverr',
    'freelancer': 'Freelancer',
    'paypal': 'Paypal',
    'payoneer': 'Payoneer',
    'binance': 'Binance',
    'macro': 'Macro',
    'físico': 'Físico', 'fisico': 'Físico',
    'cajón': 'Cajón', 'cajon': 'Cajón',
    'mercado pago': 'Mercado Pago',
    'mep': 'MEP',
    'santander': 'Santander',
    'western union': 'Western Union',
    'lucas': 'Lucas Genzelis', 'lucas genzelis': 'Lucas Genzelis',
    'previaje': 'Previaje',
    'jarras': 'Jarras',
    'saldo.com.ar': 'saldo.com.ar',
    'capitalización ah': 'Capitalización AH',
    'capitalizaci?n ah': 'Capitalización AH',
    'capitalizaci\x00n ah': 'Capitalización AH',
}

def norm_account(raw):
    if raw is None: return None
    s = str(raw).strip()
    return RAW_TO_CANONICAL.get(s.lower(), s) if s else None

# Lookup SQL para cuentas existentes en la DB, con desambiguación por moneda
# Devuelve (account_id_sql, account_name_sql)
def acc_ref(canonical, currency=None):
    if canonical is None:
        return 'NULL', 'NULL'

    def lookup(where):
        return (
            f"(SELECT id   FROM accounts WHERE {where} LIMIT 1)",
            f"(SELECT name FROM accounts WHERE {where} LIMIT 1)",
        )

    u = f"user_id = '{UID}'"

    if canonical == 'Físico':
        return lookup(f"{u} AND type = 'cash' AND currency = 'USD'")
    if canonical == 'Cajón':
        return lookup(f"{u} AND type = 'cash' AND currency = 'ARS'")
    if canonical == 'MEP':
        return lookup(f"{u} AND type = 'savings' AND currency = 'USD'")
    if canonical == 'Macro':
        if currency == 'USD':
            return lookup(f"{u} AND type = 'savings' AND currency = 'USD'")
        else:  # ARS o desconocido
            return lookup(f"{u} AND type = 'checking' AND currency = 'ARS' AND name ILIKE 'Banco Macro%'")
    if canonical == 'Mercado Pago':
        return lookup(f"{u} AND name = 'Mercado Pago'")
    if canonical == 'Capitalización AH':
        return lookup(f"{u} AND name = 'Reserva Mercado Pago'")
    if canonical == 'Binance':
        return lookup(f"{u} AND name = 'Binance'")

    # Cuenta nueva (Fiverr, Paypal, etc.) — solo nombre de texto, sin FK
    return 'NULL', esc(canonical)


# ── generación ────────────────────────────────────────────────────────────

lines = []
lines += [
    "-- ================================================================",
    "-- FinanzApp — Historial completo (generado de mami_bills.xlsx)",
    "-- user_id: 402ad72e-fbb4-4316-aa69-669dd4296bda",
    "-- Correr PRIMERO la migration 004.",
    "-- ================================================================",
    "",
]

# ── 1. CUENTAS NUEVAS (las que ya existen en la app no se tocan) ──────────
lines += [
    "-- ================================================================",
    "-- 1. CUENTAS HISTORICAS (is_visible=false)",
    "--    Solo las que NO existen aun: Binance, MEP, Fisico, Cajon",
    "--    ya estan en la app con otros nombres.",
    "-- ================================================================",
    "",
]
NEW_ACCOUNTS = [
    ('Fiverr',         'other',   'USD', 0),
    ('Freelancer',     'other',   'USD', 0),
    ('Paypal',         'other',   'USD', 0),
    ('Payoneer',       'other',   'USD', 0),
    ('Santander',      'savings', 'EUR', 0),
    ('Western Union',  'other',   'ARS', 0),
    ('Lucas Genzelis', 'other',   'USD', 0),
    ('saldo.com.ar',   'other',   'ARS', 0),
    ('Jarras',         'other',   'ARS', 0),
    ('Previaje',       'other',   'ARS', 522),
]
for name, atype, cur, bal in NEW_ACCOUNTS:
    lines.append(
        f"INSERT INTO accounts (user_id, name, type, currency, balance, is_visible, is_favorite)\n"
        f"SELECT '{UID}', {esc(name)}, '{atype}', '{cur}', {bal}, false, false\n"
        f"WHERE NOT EXISTS (\n"
        f"  SELECT 1 FROM accounts WHERE user_id = '{UID}' AND lower(name) = lower({esc(name)})\n"
        f");"
    )
lines.append("")

# ── 2. CATEGORIAS ─────────────────────────────────────────────────────────
lines += [
    "-- ================================================================",
    "-- 2. CATEGORIAS — todas ya existen en la DB, no se crean",
    "--    Trabajo > Changas  (income)",
    "--    Otros              (income)",
    "--    Vida general       (expense)",
    "-- ================================================================",
    "",
]

# ── 3. GANANCIAS ──────────────────────────────────────────────────────────
lines += [
    "-- ================================================================",
    "-- 3. GANANCIAS  →  income / Changas",
    "--    date=fecha_inicio  due_date=fecha_fin  amount=neto  amount_gross=bruto",
    "-- ================================================================",
    "",
]
ws = wb['ganancias']
rows_g = [r for r in ws.iter_rows(values_only=True) if any(v is not None for v in r)]
for row in rows_g[2:]:
    plataforma, cliente, tecnologia, concepto, fecha_ini, fecha_fin, monto_bruto, monto_neto, moneda, *_ = row
    if monto_neto is None:
        continue

    amount  = float(monto_neto)
    cur     = norm_currency(moneda)
    acc     = norm_account(plataforma)
    acc_id, acc_name = acc_ref(acc, currency=cur)

    notes_extra = None
    try:
        gross = float(monto_bruto) if monto_bruto is not None else None
    except (ValueError, TypeError):
        gross = None
        notes_extra = f"bruto original: {monto_bruto}"

    gross_sql  = fmt_num(gross) if gross is not None and abs(gross - amount) > 0.001 else 'NULL'
    notes_sql  = esc(notes_extra)
    cli_sql    = esc(str(cliente).strip() if cliente else None)
    proj_sql   = esc(str(tecnologia).strip() if tecnologia else None)
    desc_sql   = esc(str(concepto).strip() if concepto else None)

    lines.append(
        f"INSERT INTO transactions\n"
        f"  (user_id, date, due_date, type, status, amount, amount_gross, currency,\n"
        f"   description, account_id, account_name, client_name, project_name, category_name, notes)\n"
        f"VALUES (\n"
        f"  '{UID}', {fmt_date(fecha_ini)}, {fmt_date(fecha_fin)}, 'income', 'confirmed',\n"
        f"  {amount}, {gross_sql}, '{cur}',\n"
        f"  {desc_sql}, {acc_id}, {acc_name}, {cli_sql}, {proj_sql}, 'Changas', {notes_sql}\n"
        f");"
    )
lines.append("")

# ── 4. OTROS INGRESOS ─────────────────────────────────────────────────────
lines += [
    "-- ================================================================",
    "-- 4. OTROS INGRESOS  →  income (o expense si negativo) / Otros",
    "--    Sin fecha exacta (date = NULL)",
    "-- ================================================================",
    "",
]
ws = wb['otros ingresos']
rows_oi = [r for r in ws.iter_rows(values_only=True) if any(v is not None for v in r)]
for row in rows_oi[2:]:
    a_donde, concepto, monto, moneda, *_ = row
    if monto is None:
        continue

    amount  = float(monto)
    tx_type = 'expense' if amount < 0 else 'income'
    amount  = abs(amount)
    cur     = norm_currency(moneda)
    acc     = norm_account(a_donde)
    acc_id, acc_name = acc_ref(acc, currency=cur)
    desc_sql = esc(str(concepto).strip() if concepto else None)

    lines.append(
        f"INSERT INTO transactions\n"
        f"  (user_id, date, type, status, amount, currency,\n"
        f"   description, account_id, account_name, category_name)\n"
        f"VALUES (\n"
        f"  '{UID}', NULL, '{tx_type}', 'confirmed', {amount}, '{cur}',\n"
        f"  {desc_sql}, {acc_id}, {acc_name}, 'Otros'\n"
        f");"
    )
lines.append("")

# ── 5. GASTOS ─────────────────────────────────────────────────────────────
lines += [
    "-- ================================================================",
    "-- 5. GASTOS  →  expense / Vida general",
    "-- ================================================================",
    "",
]
ws = wb['gastos']
rows_gx = [r for r in ws.iter_rows(values_only=True) if any(v is not None for v in r)]
for row in rows_gx[2:]:
    concepto, medio_pago, fecha, monto, moneda, *_ = row
    if monto is None:
        continue

    amount   = float(monto)
    cur      = norm_currency(moneda)
    acc      = norm_account(medio_pago)
    acc_id, acc_name = acc_ref(acc, currency=cur)
    desc_sql = esc(str(concepto).strip() if concepto else None)

    lines.append(
        f"INSERT INTO transactions\n"
        f"  (user_id, date, type, status, amount, currency,\n"
        f"   description, account_id, account_name, category_name)\n"
        f"VALUES (\n"
        f"  '{UID}', {fmt_date(fecha)}, 'expense', 'confirmed', {amount}, '{cur}',\n"
        f"  {desc_sql}, {acc_id}, {acc_name}, 'Vida general'\n"
        f");"
    )
lines.append("")

# ── 6. TRANSFERENCIAS Y CAMBIOS ───────────────────────────────────────────
lines += [
    "-- ================================================================",
    "-- 6. TRANSFERENCIAS Y CAMBIOS",
    "--    to_amount / to_currency solo cuando hay cambio de moneda o monto distinto",
    "-- ================================================================",
    "",
]
ws = wb['transferencias y cambios']
rows_t = [r for r in ws.iter_rows(values_only=True) if any(v is not None for v in r)]
for row in rows_t[2:]:
    desde, hasta, monto_env, moneda_env, monto_rec, moneda_rec, fecha, *_ = row
    if monto_env is None:
        continue

    amount  = float(monto_env)
    cur_env = norm_currency(moneda_env)
    cur_rec = norm_currency(moneda_rec) if moneda_rec else cur_env

    acc_from = norm_account(desde)
    acc_to   = norm_account(hasta)
    # Para Macro: el que envía usa cur_env, el que recibe usa cur_rec
    src_id, src_name = acc_ref(acc_from, currency=cur_env)
    dst_id, dst_name = acc_ref(acc_to,   currency=cur_rec)

    to_amount = float(monto_rec) if monto_rec is not None else None
    is_exchange = to_amount is not None and (cur_env != cur_rec or abs(amount - to_amount) > 0.01)
    to_amount_sql   = fmt_num(to_amount) if is_exchange else 'NULL'
    to_currency_sql = esc(cur_rec)       if is_exchange and cur_rec != cur_env else 'NULL'

    lines.append(
        f"INSERT INTO transactions\n"
        f"  (user_id, date, type, status, amount, currency, to_amount, to_currency,\n"
        f"   account_id, account_name, to_account_id, to_account_name)\n"
        f"VALUES (\n"
        f"  '{UID}', {fmt_date(fecha)}, 'transfer', 'confirmed',\n"
        f"  {amount}, '{cur_env}', {to_amount_sql}, {to_currency_sql},\n"
        f"  {src_id}, {src_name}, {dst_id}, {dst_name}\n"
        f");"
    )
lines.append("")

# ── escribir ──────────────────────────────────────────────────────────────
output = '\n'.join(lines)
with open('supabase/historico.sql', 'w', encoding='utf-8') as f:
    f.write(output)

n = output.count('INSERT INTO')
print(f"OK — {n} inserts  ({output.count('INSERT INTO transactions')} transacciones, "
      f"{output.count('INSERT INTO accounts')} cuentas, "
      f"{output.count('INSERT INTO categories')} categorias)")
